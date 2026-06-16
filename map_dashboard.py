"""
百度地图 Bug 数据看板 — 高端深蓝渐变风
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from datetime import datetime
import threading
import math
import webbrowser
import os
import sys
import json
import tempfile
import subprocess
import urllib.request
import urllib.error

# ─── 颜色常量 ──────────────────────────────────────────────
BG_DARK      = "#050d1f"
BG_CARD      = "#0a1628"
BG_CARD2     = "#0f1e35"
BORDER       = "#1a3a5c"
BORDER_GLOW  = "#1e4976"
ACCENT1      = "#3b82f6"   # 蓝
ACCENT2      = "#06b6d4"   # 青
ACCENT3      = "#f59e0b"   # 金
ACCENT_RED   = "#ef4444"
ACCENT_GREEN = "#22c55e"
TEXT_H       = "#e2eaf5"
TEXT_M       = "#8ba3c0"
TEXT_D       = "#3d5a7a"
ROW_ODD      = "#0c1a2e"
ROW_EVEN     = "#091525"
ROW_HOVER    = "#122034"
SEL_BG       = "#1a3a5c"
HEADER_BG    = "#071020"
SCROLLBAR_BG = "#0a1628"

FONT_TITLE  = ("微软雅黑", 22, "bold")
FONT_SUB    = ("微软雅黑", 11)
FONT_LABEL  = ("微软雅黑", 9)
FONT_DATA   = ("Consolas", 9)
FONT_BADGE  = ("微软雅黑", 9, "bold")
FONT_TAB    = ("微软雅黑", 10, "bold")
FONT_STAT   = ("微软雅黑", 28, "bold")
FONT_STAT_L = ("微软雅黑", 13)

APP_VERSION = "v1.0.0"
GITHUB_RELEASES_LATEST = "https://github.com/1250602390/POITab/releases/latest"
GITHUB_RELEASES_DOWNLOAD = "https://github.com/1250602390/POITab/releases/latest/download/POI%E6%A0%87%E7%AD%BE%E5%88%86%E7%B1%BB.exe"
APP_EXE_NAME = "POI标签分类.exe"


# ─── 自动更新 ──────────────────────────────────────────────
def parse_version(v):
    """将 v1.2.3 转为 (1,2,3)，无法识别时返回 (0,)"""
    if not v:
        return (0,)
    v = str(v).strip().lstrip("vV")
    parts = []
    for part in v.split("."):
        digits = "".join(ch for ch in part if ch.isdigit())
        parts.append(int(digits or 0))
    return tuple(parts or [0])

def is_newer_version(remote, local):
    r = parse_version(remote)
    l = parse_version(local)
    max_len = max(len(r), len(l))
    r += (0,) * (max_len - len(r))
    l += (0,) * (max_len - len(l))
    return r > l

def get_current_exe_path():
    if getattr(sys, "frozen", False):
        return sys.executable
    return os.path.abspath(__file__)

# ─── 数据加载 ──────────────────────────────────────────────
def load_excel(path):
    # 不能使用 read_only=True，否则单元格 hyperlink 信息会丢失
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    ws = wb.active
    if ws.max_row < 1:
        wb.close()
        return [], []

    header_cells = list(ws[1])
    headers = [cell.value for cell in header_cells]
    title_col = None
    for cell in header_cells:
        if cell.value == "标题":
            title_col = cell.column
            break

    data = []
    for row_cells in ws.iter_rows(min_row=2):
        values = tuple(cell.value for cell in row_cells)
        title_link = ""
        if title_col:
            cell = row_cells[title_col - 1]
            if cell.hyperlink:
                title_link = cell.hyperlink.target or cell.hyperlink.location or ""
        data.append({"values": values, "title_link": title_link})
    wb.close()
    return headers, data

def find_col(headers, name):
    for i, h in enumerate(headers):
        if h == name:
            return i
    for i, h in enumerate(headers):
        if h and name in str(h):
            return i
    return None

def fmt_dt(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M")
    return str(v) if v else ""

def fmt_dur(hours):
    if hours is None:
        return "—"
    if hours < 0:
        return "数据异常"
    h = int(hours)
    m = int((hours - h) * 60)
    if h >= 24:
        d = h // 24
        rh = h % 24
        return f"{d}天{rh}h{m:02d}m"
    return f"{h}h{m:02d}m"

def build_rows(headers, data, source_file=""):
    c = {
        "标题":      find_col(headers, "标题"),
        "创建时间":  find_col(headers, "创建时间"),
        "解决时间":  find_col(headers, "解决时间"),
        "POI-处理人":find_col(headers, "POI-处理人"),
        "问题定位信息": find_col(headers, "问题定位信息"),
        "流程状态":  find_col(headers, "流程状态"),
    }
    rows = []
    for item in data:
        r = item["values"] if isinstance(item, dict) else item
        title_link = item.get("title_link", "") if isinstance(item, dict) else ""
        create = r[c["创建时间"]] if c["创建时间"] is not None else None
        solve  = r[c["解决时间"]]  if c["解决时间"]  is not None else None
        hours  = None
        if isinstance(create, datetime) and isinstance(solve, datetime):
            hours = (solve - create).total_seconds() / 3600
        rows.append({
            "标题":       str(r[c["标题"]] or "") if c["标题"]       is not None else "",
            "标题链接":   title_link,
            "创建时间":   fmt_dt(create),
            "解决时间":   fmt_dt(solve),
            "POI-处理人": str(r[c["POI-处理人"]] or "") if c["POI-处理人"] is not None else "",
            "问题定位信息": str(r[c["问题定位信息"]] or "") if c["问题定位信息"] is not None else "",
            "流程状态":   str(r[c["流程状态"]] or "") if c["流程状态"] is not None else "",
            "完成时间":   fmt_dur(hours),
            "完成时间_h": hours,
            "创建时间_dt": create,
            "解决时间_dt": solve,
            "来源文件":   source_file,
        })
    return rows

CLOSED_STATES = {
    # excel-export (58) 状态
    "Resolved（已完成上线/发布）",
    "无效Bug（使用错误/不合理建议）",
    "重复case",
    "需求建议（PM跟进）",
    "case积累",
    # excel-export (59) 状态
    "已完成",
    "无效",
}

def calc_stats(rows):
    total = len(rows)
    closed = sum(1 for r in rows if r["流程状态"] in CLOSED_STATES)
    no_solve = sum(1 for r in rows if r["完成时间"] == "—")
    over24 = sum(1 for r in rows if r["完成时间_h"] is not None and r["完成时间_h"] > 24)
    in24   = sum(1 for r in rows if r["完成时间_h"] is not None and r["完成时间_h"] <= 24)
    urgent_rows = [r for r in rows if "急" in str(r.get("标题", ""))]
    urgent_in1 = sum(1 for r in urgent_rows if r.get("完成时间_h") is not None and r["完成时间_h"] <= 1)
    urgent_over1 = len(get_urgent_over1(rows))
    rate   = closed / total * 100 if total else 0
    rate_24h = in24 / total * 100 if total else 0
    urgent_rate_1h = urgent_in1 / len(urgent_rows) * 100 if urgent_rows else 0
    return dict(total=total, closed=closed, no_solve=no_solve,
                over24=over24, in24=in24, urgent_total=len(urgent_rows),
                urgent_in1=urgent_in1, urgent_over1=urgent_over1,
                rate=rate, rate_24h=rate_24h, urgent_rate_1h=urgent_rate_1h)

def get_unclosed(rows):
    return [r for r in rows if r["流程状态"] not in CLOSED_STATES]

def get_over24(rows):
    return [r for r in rows if r["完成时间_h"] is not None and r["完成时间_h"] > 24]

def get_urgent_over1(rows):
    """标题含“急”，且完成时间超过1小时；未填写解决时间时按当前时间估算。"""
    result = []
    now = datetime.now()
    for r in rows:
        if "急" not in str(r.get("标题", "")):
            continue
        hours = r.get("完成时间_h")
        if hours is None:
            create = r.get("创建时间_dt")
            if isinstance(create, datetime):
                hours = (now - create).total_seconds() / 3600
                if hours > 1:
                    rr = dict(r)
                    rr["完成时间"] = "未解决 / " + fmt_dur(hours)
                    rr["完成时间_h"] = hours
                    result.append(rr)
            continue
        if hours > 1:
            result.append(r)
    return result


# ─── 圆弧进度条（Canvas）─────────────────────────────────
class ArcGauge(tk.Canvas):
    def __init__(self, parent, size=160, label="闭环率", **kw):
        super().__init__(parent, width=size, height=size,
                         bg=BG_CARD, highlightthickness=0, **kw)
        self._size = size
        self._label = label
        self._rate = 0
        self._target = 0
        self._draw(0)

    def _draw(self, rate):
        s = self._size
        m = 14
        self.delete("all")
        # 背景轨道
        self.create_arc(m, m, s-m, s-m, start=0, extent=360,
                        outline=BG_CARD2, width=10, style="arc")
        # 进度弧
        if rate > 0:
            extent = rate / 100 * 360
            color = ACCENT_GREEN if rate >= 90 else ACCENT1 if rate >= 70 else ACCENT3
            self.create_arc(m, m, s-m, s-m, start=90, extent=-extent,
                            outline=color, width=10, style="arc")
        # 中央只显示百分比数字
        cx, cy = s // 2, s // 2
        self.create_text(cx, cy, text=f"{rate:.1f}%",
                         font=("微软雅黑", 22, "bold"), fill=TEXT_H)

    def animate_to(self, target):
        self._target = target
        self._rate = 0
        self._step()

    def _step(self):
        if self._rate < self._target:
            self._rate = min(self._rate + 1.5, self._target)
            self._draw(self._rate)
            self.after(12, self._step)
        else:
            self._draw(self._target)


# ─── 自定义表格组件 ───────────────────────────────────────
class StyledTable(tk.Frame):
    COLS = {
        "标题":       ("标题", 320),
        "创建时间":   ("创建时间", 130),
        "解决时间":   ("解决时间", 130),
        "完成时间":   ("完成时间", 90),
        "POI-处理人": ("POI-处理人", 120),
        "问题定位信息":("问题定位信息", 220),
        "流程状态":   ("流程状态", 140),
        "来源文件":   ("来源文件", 180),
    }

    def __init__(self, parent, col_keys, tag_col=None, **kw):
        super().__init__(parent, bg=BG_DARK, **kw)
        self._tag_col = tag_col

        style = ttk.Style()
        style.theme_use("clam")
        uid = f"styled{id(self)}"
        style.configure(f"{uid}.Treeview",
            background=ROW_ODD, foreground=TEXT_H,
            fieldbackground=ROW_ODD, rowheight=28,
            borderwidth=0, font=FONT_DATA)
        style.configure(f"{uid}.Treeview.Heading",
            background=HEADER_BG, foreground=ACCENT2,
            relief="flat", font=("微软雅黑", 9, "bold"), borderwidth=0)
        style.map(f"{uid}.Treeview",
            background=[("selected", SEL_BG)],
            foreground=[("selected", TEXT_H)])

        cols = [k for k in col_keys if k in self.COLS]
        self._tree = ttk.Treeview(self, columns=cols, show="headings",
                                   style=f"{uid}.Treeview")
        for k in cols:
            label, width = self.COLS[k]
            self._tree.heading(k, text=label)
            self._tree.column(k, width=width, minwidth=60, anchor="w")

        sb_y = tk.Scrollbar(self, orient="vertical", command=self._tree.yview,
                             bg=SCROLLBAR_BG, troughcolor=BG_DARK,
                             activebackground=ACCENT1, relief="flat", width=8)
        sb_x = tk.Scrollbar(self, orient="horizontal", command=self._tree.xview,
                             bg=SCROLLBAR_BG, troughcolor=BG_DARK,
                             activebackground=ACCENT1, relief="flat", width=8)
        self._tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        sb_y.grid(row=0, column=1, sticky="ns")
        sb_x.grid(row=1, column=0, sticky="ew")
        self.rowconfigure(0, weight=1)
        self.columnconfigure(0, weight=1)

        self._tree.tag_configure("odd",  background=ROW_ODD,  foreground=TEXT_H)
        self._tree.tag_configure("even", background=ROW_EVEN, foreground=TEXT_H)
        self._tree.tag_configure("warn", background="#1a1a08", foreground=ACCENT3)
        self._tree.tag_configure("danger", background="#1a0808", foreground=ACCENT_RED)
        self._tree.tag_configure("link", foreground=ACCENT2)

        # 存储 iid -> 链接的映射
        self._links = {}
        self._tree.bind("<ButtonRelease-1>", self._on_click)
        self._tree.bind("<Motion>", self._on_motion)

        self._col_keys = cols

    def _on_click(self, event):
        item = self._tree.identify_row(event.y)
        col_id = self._tree.identify_column(event.x)
        if not item or not col_id:
            return
        col_index = int(col_id.replace("#", "")) - 1
        if col_index < len(self._col_keys) and self._col_keys[col_index] == "标题":
            url = self._links.get(item)
            if url:
                webbrowser.open(url)

    def _on_motion(self, event):
        item = self._tree.identify_row(event.y)
        col_id = self._tree.identify_column(event.x)
        if item and col_id:
            col_index = int(col_id.replace("#", "")) - 1
            if col_index < len(self._col_keys) and self._col_keys[col_index] == "标题":
                if self._links.get(item):
                    self._tree.configure(cursor="hand2")
                    return
        self._tree.configure(cursor="")

    def load(self, rows):
        self._tree.delete(*self._tree.get_children())
        self._links = {}
        for i, r in enumerate(rows):
            values = [r.get(k, "") for k in self._col_keys]
            # 颜色规则：红色=标题含“急”且超过1小时；黄色=超过24小时
            tag = "odd" if i % 2 == 0 else "even"
            h = r.get("完成时间_h")
            title = str(r.get("标题", ""))
            urgent_over1 = False
            if "急" in title:
                if h is not None:
                    urgent_over1 = h > 1
                else:
                    create = r.get("创建时间_dt")
                    if isinstance(create, datetime):
                        urgent_over1 = (datetime.now() - create).total_seconds() / 3600 > 1
            if urgent_over1:
                tag = "danger"
            elif h is not None and h > 24:
                tag = "warn"
            iid = self._tree.insert("", "end", values=values, tags=(tag,))
            if r.get("标题链接"):
                self._links[iid] = r.get("标题链接")


# ─── 统计卡片 ──────────────────────────────────────────────
class StatCard(tk.Frame):
    def __init__(self, parent, label, value, color=ACCENT1, **kw):
        super().__init__(parent, bg=BG_CARD2,
                         highlightbackground=BORDER, highlightthickness=1, **kw)
        self.value_label = tk.Label(self, text=value, font=FONT_STAT, fg=color, bg=BG_CARD2)
        self.value_label.pack(pady=(18,2))
        tk.Label(self, text=label, font=FONT_STAT_L, fg=TEXT_M, bg=BG_CARD2).pack(pady=(0,16))

    def set_value(self, value):
        self.value_label.configure(text=str(value))


# ─── 主应用 ───────────────────────────────────────────────
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("POI标签分类")
        self.geometry("1440x860")
        self.minsize(1100, 700)
        self.configure(bg=BG_DARK)

        self._rows = []
        self._selected_paths = []
        self._file_path = tk.StringVar(value="未加载文件")
        self._build_ui()
        # 启动后稍等再联网检测，避免影响界面打开速度
        self.after(1500, self._check_update_async)

    def _check_update_async(self):
        threading.Thread(target=self._check_update_thread, daemon=True).start()

    def _check_update_thread(self):
        # 仅在打包后的 EXE 中执行自动更新；源码调试时跳过，避免误替换 .py 文件
        if not getattr(sys, "frozen", False):
            return
        try:
            # 不使用 GitHub REST API，避免未认证 API rate limit。
            # 访问 releases/latest 会被 GitHub 重定向到 /releases/tag/vX.Y.Z，直接从最终 URL 提取版本号。
            req = urllib.request.Request(
                GITHUB_RELEASES_LATEST,
                headers={"User-Agent": f"POITab-Updater/{APP_VERSION}"},
                method="HEAD"
            )
            try:
                with urllib.request.urlopen(req, timeout=10) as resp:
                    final_url = resp.geturl()
            except Exception:
                # 某些网络代理不支持 HEAD，退回 GET，但不读取正文。
                req = urllib.request.Request(
                    GITHUB_RELEASES_LATEST,
                    headers={"User-Agent": f"POITab-Updater/{APP_VERSION}"}
                )
                with urllib.request.urlopen(req, timeout=10) as resp:
                    final_url = resp.geturl()

            latest_version = final_url.rstrip("/").split("/")[-1]
            if not latest_version or latest_version == "latest":
                return
            if not is_newer_version(latest_version, APP_VERSION):
                self.after(0, lambda: self._statusbar.configure(
                    text=f"当前已是最新版本 {APP_VERSION}"))
                return

            self.after(0, lambda: self._prompt_and_update(
                latest_version, APP_EXE_NAME, GITHUB_RELEASES_DOWNLOAD))
        except Exception:
            # 更新检测失败不影响主功能，静默跳过
            return

    def _prompt_and_update(self, latest_version, asset_name, asset_url):
        # 自动下载并替换：用户已要求自动更新，这里不再二次确认
        self._statusbar.configure(
            text=f"发现新版本 {latest_version}，正在自动下载并准备替换：{asset_name}")
        threading.Thread(target=self._download_and_replace,
                         args=(latest_version, asset_url), daemon=True).start()

    def _download_and_replace(self, latest_version, asset_url):
        try:
            temp_dir = tempfile.gettempdir()
            new_exe = os.path.join(temp_dir, f"POI标签分类_{latest_version}.exe")
            req = urllib.request.Request(
                asset_url,
                headers={"User-Agent": f"POITab-Updater/{APP_VERSION}"}
            )
            with urllib.request.urlopen(req, timeout=60) as resp, open(new_exe, "wb") as f:
                while True:
                    chunk = resp.read(1024 * 256)
                    if not chunk:
                        break
                    f.write(chunk)

            current_exe = get_current_exe_path()
            bat_path = os.path.join(temp_dir, "POI标签分类_自动更新.bat")
            bat = f'''@echo off
chcp 65001 >nul
set "TARGET={current_exe}"
set "NEWEXE={new_exe}"
timeout /t 2 /nobreak >nul
:retry
del /f /q "%TARGET%" >nul 2>nul
if exist "%TARGET%" (
  timeout /t 1 /nobreak >nul
  goto retry
)
move /Y "%NEWEXE%" "%TARGET%" >nul
start "" "%TARGET%"
del "%~f0"
'''
            with open(bat_path, "w", encoding="utf-8") as f:
                f.write(bat)

            self.after(0, lambda: self._statusbar.configure(text="下载完成，正在重启更新…"))
            os.startfile(bat_path)
            self.after(300, self.destroy)
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("自动更新失败", str(e)))

    def _build_ui(self):
        # ── 顶部标题栏 ──
        hdr = tk.Frame(self, bg=BG_CARD, height=64)
        hdr.pack(fill="x", side="top")
        hdr.pack_propagate(False)

        tk.Label(hdr, text="🗺  POI标签分类",
                 font=FONT_TITLE, fg=TEXT_H, bg=BG_CARD).pack(side="left", padx=28, pady=10)

        # 文件选择区
        fr = tk.Frame(hdr, bg=BG_CARD)
        fr.pack(side="right", padx=20)
        self._file_count_label = tk.Label(fr, textvariable=self._file_path,
                 font=FONT_LABEL, fg=TEXT_M, bg=BG_CARD)
        self._file_count_label.pack(side="left", padx=8)
        tk.Button(fr, text="  ✖ 清空  ", command=self._clear_files,
                  font=FONT_BADGE, fg=TEXT_M, bg=BG_CARD2,
                  activebackground=ACCENT_RED, activeforeground=TEXT_H,
                  relief="flat", cursor="hand2", padx=8, pady=6).pack(side="left", padx=4)
        tk.Button(fr, text="  📂 添加 Excel  ", command=self._pick_file,
                  font=FONT_BADGE, fg=TEXT_H, bg=ACCENT1,
                  activebackground=ACCENT2, activeforeground=TEXT_H,
                  relief="flat", cursor="hand2", padx=10, pady=6).pack(side="left")

        # 分隔线
        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        # ── Tab 区 ──
        self._tab_frame = tk.Frame(self, bg=BG_DARK)
        self._tab_frame.pack(fill="x", padx=0, pady=0)

        self._tabs = {}
        self._active_tab = tk.StringVar(value="数据总览")
        tab_names = ["数据总览", "闭环率分析", "未闭环 & 超时"]
        for name in tab_names:
            b = tk.Button(self._tab_frame, text=name,
                          font=FONT_TAB, relief="flat",
                          cursor="hand2", pady=10, padx=24,
                          command=lambda n=name: self._switch_tab(n))
            b.pack(side="left")
            self._tabs[name] = b

        tk.Frame(self, bg=BORDER, height=1).pack(fill="x")

        # ── 内容区 ──
        self._content = tk.Frame(self, bg=BG_DARK)
        self._content.pack(fill="both", expand=True)

        self._pages = {}
        self._pages["数据总览"]    = self._build_page_overview()
        self._pages["闭环率分析"]  = self._build_page_rate()
        self._pages["未闭环 & 超时"] = self._build_page_unclosed()

        self._switch_tab("数据总览")

        # 状态栏
        self._statusbar = tk.Label(self, text="请点击右上角 [选择 Excel] 加载数据",
                                   font=FONT_LABEL, fg=TEXT_M, bg=BG_CARD,
                                   anchor="w", padx=14)
        self._statusbar.pack(fill="x", side="bottom")

    def _switch_tab(self, name):
        self._active_tab.set(name)
        for n, b in self._tabs.items():
            if n == name:
                b.configure(bg=ACCENT1, fg=TEXT_H)
            else:
                b.configure(bg=BG_DARK, fg=TEXT_M)
        for n, p in self._pages.items():
            if n == name:
                p.pack(fill="both", expand=True)
            else:
                p.pack_forget()

    # ── 总览页 ──────────────────────────────────
    def _build_page_overview(self):
        page = tk.Frame(self._content, bg=BG_DARK)
        # 工具栏
        bar = tk.Frame(page, bg=BG_DARK)
        bar.pack(fill="x", padx=16, pady=(10, 4))
        tk.Label(bar, text="数据字段总览", font=("微软雅黑", 12, "bold"),
                 fg=TEXT_H, bg=BG_DARK).pack(side="left")
        self._ov_count = tk.Label(bar, text="共 0 条", font=FONT_LABEL,
                                   fg=TEXT_M, bg=BG_DARK)
        self._ov_count.pack(side="left", padx=10)
        # 表格
        col_keys = ["标题", "创建时间", "解决时间", "完成时间", "POI-处理人", "问题定位信息", "流程状态", "来源文件"]
        self._ov_table = StyledTable(page, col_keys)
        self._ov_table.pack(fill="both", expand=True, padx=16, pady=(0, 10))
        return page

    # ── 闭环率页 ─────────────────────────────────
    def _build_page_rate(self):
        page = tk.Frame(self._content, bg=BG_DARK)

        # ── 第一行：三个仪表盘横排 ──
        gauge_row = tk.Frame(page, bg=BG_DARK)
        gauge_row.pack(fill="x", padx=24, pady=(16, 0))

        self._gauges = {}
        gauge_defs = [
            ("rate",          "整体闭环率",   "按流程状态"),
            ("rate_24h",      "24小时闭环率", "24h完成 / 总记录"),
            ("urgent_rate_1h","急单1小时闭环率","急单1h完成 / 急单总数"),
        ]
        for i, (key, title, sub) in enumerate(gauge_defs):
            card = tk.Frame(gauge_row, bg=BG_CARD,
                            highlightbackground=BORDER, highlightthickness=1)
            card.grid(row=0, column=i, padx=(0 if i == 0 else 12, 0), sticky="nsew")
            gauge_row.columnconfigure(i, weight=1)
            tk.Label(card, text=title, font=("微软雅黑", 11, "bold"),
                     fg=ACCENT2, bg=BG_CARD).pack(pady=(12, 0))
            tk.Label(card, text=sub, font=FONT_LABEL,
                     fg=TEXT_D, bg=BG_CARD).pack(pady=(0, 4))
            g = ArcGauge(card, size=150, label=title)
            g.pack(padx=24, pady=(0, 14))
            self._gauges[key] = g

        # ── 第二行：统计卡片横排（分两组，用分隔线区分）──
        stats_row = tk.Frame(page, bg=BG_DARK)
        stats_row.pack(fill="x", padx=24, pady=(14, 0))

        self._stat_cards = {}
        # 左组：整体
        left_grp = tk.Frame(stats_row, bg=BG_CARD2,
                             highlightbackground=BORDER, highlightthickness=1)
        left_grp.pack(side="left", fill="both", expand=True, padx=(0, 8))
        tk.Label(left_grp, text="整体数据", font=FONT_BADGE, fg=TEXT_M,
                 bg=BG_CARD2).pack(anchor="w", padx=14, pady=(8, 4))
        left_cards = tk.Frame(left_grp, bg=BG_CARD2)
        left_cards.pack(fill="x", padx=8, pady=(0, 10))
        for i, (key, label, color) in enumerate([
            ("total",  "总记录数",     ACCENT2),
            ("closed", "已闭环",       ACCENT_GREEN),
            ("in24",   "24h内完成",    ACCENT1),
            ("over24", "超24小时",     ACCENT3),
        ]):
            c = StatCard(left_cards, label, "—", color)
            c.grid(row=0, column=i, padx=4, pady=0, sticky="nsew")
            left_cards.columnconfigure(i, weight=1)
            self._stat_cards[key] = c

        # 右组：急单
        right_grp = tk.Frame(stats_row, bg=BG_CARD2,
                              highlightbackground=BORDER, highlightthickness=1)
        right_grp.pack(side="left", fill="both", expand=True)
        tk.Label(right_grp, text='急单数据（标题含"急"）', font=FONT_BADGE, fg=TEXT_M,
                 bg=BG_CARD2).pack(anchor="w", padx=14, pady=(8, 4))
        right_cards = tk.Frame(right_grp, bg=BG_CARD2)
        right_cards.pack(fill="x", padx=8, pady=(0, 10))
        for i, (key, label, color) in enumerate([
            ("urgent_total", "急单总数",     TEXT_M),
            ("urgent_in1",   "1h内完成",     ACCENT_GREEN),
            ("urgent_over1", "超1小时",      ACCENT_RED),
        ]):
            c = StatCard(right_cards, label, "—", color)
            c.grid(row=0, column=i, padx=4, pady=0, sticky="nsew")
            right_cards.columnconfigure(i, weight=1)
            self._stat_cards[key] = c

        # ── 第三行：状态分布条 ──
        tk.Frame(page, bg=BORDER, height=1).pack(fill="x", padx=24, pady=(14, 0))
        bot = tk.Frame(page, bg=BG_DARK)
        bot.pack(fill="both", expand=True, padx=24, pady=(10, 16))
        tk.Label(bot, text="流程状态分布", font=("微软雅黑", 11, "bold"),
                 fg=TEXT_H, bg=BG_DARK).pack(anchor="w", pady=(0, 8))
        self._status_frame = tk.Frame(bot, bg=BG_DARK)
        self._status_frame.pack(fill="both", expand=True)
        return page

    # ── 未闭环&超时页 ───────────────────────────
    def _build_page_unclosed(self):
        page = tk.Frame(self._content, bg=BG_DARK)

        nb = tk.Frame(page, bg=BG_DARK)
        nb.pack(fill="x", padx=16, pady=(10, 4))

        self._uc_tabs = {}
        self._uc_active = tk.StringVar(value="未闭环")
        for name in ["未闭环", "超24小时", "急单超1小时"]:
            b = tk.Button(nb, text=name, font=("微软雅黑", 9, "bold"),
                          relief="flat", cursor="hand2", pady=6, padx=16,
                          command=lambda n=name: self._switch_uc(n))
            b.pack(side="left", padx=2)
            self._uc_tabs[name] = b

        self._uc_count = tk.Label(nb, text="共 0 条", font=FONT_LABEL,
                                   fg=TEXT_M, bg=BG_DARK)
        self._uc_count.pack(side="left", padx=8)

        col_keys = ["标题", "创建时间", "解决时间", "完成时间", "POI-处理人", "问题定位信息", "流程状态", "来源文件"]
        self._uc_table = StyledTable(page, col_keys)
        self._uc_table.pack(fill="both", expand=True, padx=16, pady=(0, 10))

        self._switch_uc("未闭环")
        return page

    def _switch_uc(self, name):
        self._uc_active.set(name)
        for n, b in self._uc_tabs.items():
            if n == name:
                b.configure(bg=ACCENT1, fg=TEXT_H)
            else:
                b.configure(bg=BG_CARD, fg=TEXT_M)
        if not self._rows:
            return
        if name == "未闭环":
            data = get_unclosed(self._rows)
        elif name == "超24小时":
            data = get_over24(self._rows)
        else:
            data = get_urgent_over1(self._rows)
        self._uc_count.configure(text=f"共 {len(data)} 条")
        self._uc_table.load(data)

    # ── 选文件 ───────────────────────────────────
    def _pick_file(self):
        paths = filedialog.askopenfilenames(
            title="选择一个或多个 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if not paths:
            return

        existed = set(self._selected_paths)
        added = [p for p in paths if p not in existed]
        if not added:
            self._statusbar.configure(text="选择的文件已全部加载，无需重复添加")
            return

        self._selected_paths.extend(added)
        self._update_file_label()
        self._statusbar.configure(text=f"正在加载 {len(self._selected_paths)} 个 Excel 文件…")
        threading.Thread(target=self._load_thread, args=(list(self._selected_paths),), daemon=True).start()

    def _clear_files(self):
        self._selected_paths = []
        self._rows = []
        self._file_path.set("未加载文件")
        self._ov_count.configure(text="共 0 条")
        self._ov_table.load([])
        for g in self._gauges.values():
            g.animate_to(0)
        for sc in self._stat_cards.values():
            sc.set_value("—")
        self._render_status_bars([])
        self._uc_count.configure(text="共 0 条")
        self._uc_table.load([])
        self._statusbar.configure(text="已清空，请重新添加 Excel 文件")

    def _update_file_label(self):
        count = len(self._selected_paths)
        if count == 0:
            self._file_path.set("未加载文件")
        elif count == 1:
            name = self._selected_paths[0].split("/")[-1].split("\\")[-1]
            self._file_path.set(f"📄 {name}")
        else:
            self._file_path.set(f"📚 已加载 {count} 个 Excel 文件")

    def _load_thread(self, paths):
        try:
            all_rows = []
            errors = []
            for path in paths:
                name = path.split("/")[-1].split("\\")[-1]
                try:
                    headers, data = load_excel(path)
                    all_rows.extend(build_rows(headers, data, source_file=name))
                except Exception as e:
                    errors.append(f"{name}: {e}")
            self.after(0, lambda: self._on_loaded(all_rows, errors))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("加载失败", str(e)))

    def _on_loaded(self, rows, errors=None):
        errors = errors or []
        self._rows = rows
        stats = calc_stats(rows)
        # 总览
        self._ov_count.configure(text=f"共 {len(rows)} 条")
        self._ov_table.load(rows)
        # 闭环率
        for key, gauge in self._gauges.items():
            gauge.animate_to(stats[key])
        for key, sc in self._stat_cards.items():
            sc.set_value(stats[key])
        self._render_status_bars(rows)
        # 未闭环
        self._switch_uc(self._uc_active.get())
        file_count = len(self._selected_paths)
        err_text = f"  |  {len(errors)} 个文件加载失败" if errors else ""
        self._statusbar.configure(
            text=f"✅ 已加载 {file_count} 个文件，共 {len(rows)} 条数据  |  闭环率 {stats['rate']:.1f}%  "
                 f"|  超24小时 {stats['over24']} 条  |  急单超1小时 {stats['urgent_over1']} 条  "
                 f"|  未闭环 {len(get_unclosed(rows))} 条{err_text}")
        if errors:
            messagebox.showwarning("部分文件加载失败", "\n".join(errors[:8]))

    def _render_status_bars(self, rows):
        for w in self._status_frame.winfo_children():
            w.destroy()
        from collections import Counter
        counts = Counter(r["流程状态"] for r in rows)
        total = len(rows)
        if total == 0:
            tk.Label(self._status_frame, text="暂无数据", font=FONT_LABEL,
                     fg=TEXT_M, bg=BG_DARK).pack(anchor="w", pady=8)
            return
        colors = {
            "Resolved（已完成上线/发布）": ACCENT_GREEN,
            "已完成": ACCENT_GREEN,
            "无效Bug（使用错误/不合理建议）": TEXT_M,
            "无效": TEXT_M,
            "待RD/数据修复": ACCENT3,
            "重复case": ACCENT2,
        }
        for i, (st, cnt) in enumerate(counts.most_common()):
            pct = cnt / total * 100
            color = colors.get(st, ACCENT1)
            row_fr = tk.Frame(self._status_frame, bg=BG_DARK)
            row_fr.pack(fill="x", pady=3)
            tk.Label(row_fr, text=st, font=FONT_LABEL, fg=TEXT_M, bg=BG_DARK,
                     width=26, anchor="w").pack(side="left")
            bar_bg = tk.Frame(row_fr, bg=BG_CARD2, height=14)
            bar_bg.pack(side="left", fill="x", expand=True, padx=8)
            bar_bg.update_idletasks()
            bar_fill = tk.Frame(bar_bg, bg=color, height=14)
            bar_fill.place(x=0, y=0, relwidth=min(pct / 100, 1.0), relheight=1.0)
            tk.Label(row_fr, text=f"{cnt}  ({pct:.1f}%)", font=FONT_LABEL,
                     fg=TEXT_H, bg=BG_DARK, width=14, anchor="w").pack(side="left")


if __name__ == "__main__":
    app = App()
    app.mainloop()

